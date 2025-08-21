import io
import os
import re
from datetime import datetime

import pandas as pd
from django.contrib import messages
from django.http import HttpResponse
from django.shortcuts import render, redirect
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill

def bloc3(request):
    return render(request, 'bloc3.html', {"user_name": "Utilisateur"})


def export_block3(request):
    if request.method == 'POST' and request.FILES.get('file_block3'):
        uploaded_file = request.FILES['file_block3']

        if not uploaded_file.name.lower().endswith('.txt'):
            messages.error(request, 'Veuillez sélectionner un fichier .txt valide.')
            return redirect('bloc3')

        try:
            # Lire et parser les messages WhatsApp
            content = uploaded_file.read().decode('utf-8', errors='ignore')
            lines = content.splitlines()
            pattern = r"(\d{2}/\d{2}/\d{4}),\s(\d{2}:\d{2})\s-\s([^:]+):\s(.*)"

            records = []
            for line in lines:
                m = re.match(pattern, line.strip())
                if m:
                    records.append(m.groups())

            if not records:
                messages.error(request, "Aucune donnée valide trouvée.")
                return redirect('bloc3')

            cols = ['Date', 'Heure', 'Auteur', 'Message']
            df = pd.DataFrame(records, columns=cols)
            df['Date_heure'] = pd.to_datetime(df['Date'] + ' ' + df['Heure'], dayfirst=True, errors='coerce')
            df = df.sort_values('Date_heure').reset_index(drop=True)
            df.drop(columns=['Date_heure'], inplace=True)

            # Feuille 1: messages avec formatage simple (masquer répétitions Auteur/Date)
            feuille1_rows = []
            prev_date = None
            prev_author = None
            for _, row in df.iterrows():
                r = row.to_dict()
                if prev_date == row['Date'] and prev_author == row['Auteur']:
                    r['Auteur'] = ''
                    r['Date'] = ''
                else:
                    prev_date = row['Date']
                    prev_author = row['Auteur']
                feuille1_rows.append(r)
            df_feuille1 = pd.DataFrame(feuille1_rows, columns=cols)[['Auteur', 'Date', 'Heure', 'Message']]

            # Feuille 2: messages regroupés par Date/Auteur avec séparations
            df['Date_dt'] = pd.to_datetime(df['Date'], dayfirst=True, errors='coerce')
            df_sorted = df.sort_values(['Date_dt', 'Auteur', 'Heure']).reset_index(drop=True)
            grouped_rows = []
            prev_date = None
            prev_author = None
            for _, row in df_sorted.iterrows():
                current_date = row['Date_dt']
                current_author = row['Auteur']
                if prev_date is not None and current_date != prev_date:
                    for _ in range(3):
                        grouped_rows.append({'Date': '', 'Heure': '', 'Auteur': '', 'Message': ''})
                new_row = row.to_dict()
                new_row['Date'] = '' if prev_date == current_date else row['Date']
                prev_date = current_date
                new_row['Auteur'] = '' if prev_author == current_author and new_row['Date'] == '' else row['Auteur']
                prev_author = current_author
                grouped_rows.append(new_row)
            df_feuille2 = pd.DataFrame(grouped_rows, columns=cols)[['Date', 'Auteur', 'Heure', 'Message']]

            # Feuille 3/4: colonnes prédéfinies (comme l'ancien projet)
            left_columns = ['Concatener', 'DATE', 'ZI', 'TECHNICIEN', 'NOTE', 'HEURE DE TRAVAIL']
            left_columns_feuille4 = ['Concatener', 'DATE', 'ZI', 'TECHNICIEN', 'NOK', 'OK', 'SOMME', 'NOTE', 'HEURE DE TRAVAIL']
            driving_headers = ['compteur n départ', 'Localisation', 'Heure de départ']
            site_headers_site1 = ['Code site', 'Localisation', 'Heure début', 'Heure fin', 'Localisation']
            site_headers_site2 = ['Code site', 'Localisation', 'Heure début', 'Heure fin', 'Localisation']
            feuille3_columns = left_columns + driving_headers + site_headers_site1 + site_headers_site2
            feuille4_columns = left_columns_feuille4 + driving_headers + site_headers_site1 + site_headers_site2

            df_unique = df[['Date', 'Auteur']].drop_duplicates().reset_index(drop=True)
            df_feuille3 = pd.DataFrame(index=range(len(df_unique)), columns=feuille3_columns)
            df_feuille4 = pd.DataFrame(index=range(len(df_unique)), columns=feuille4_columns)
            df_feuille3['DATE'] = df_unique['Date'].values
            df_feuille3['TECHNICIEN'] = df_unique['Auteur'].values
            df_feuille4['DATE'] = df_unique['Date'].values
            df_feuille4['TECHNICIEN'] = df_unique['Auteur'].values

            def extract_city_from_filename(filename: str) -> str:
                base = os.path.splitext(filename)[0]
                tokens = re.split(r'[\s_\-]+', base)
                ignore = {'perso', 'bloc', 'bloc3', 'export', 'whatsapp', 'messages', 'chat', 'data', 'result', 'ocr', 'pdf', 'zip', 'txt', 'xlsx', 'xls'}
                for token in tokens:
                    if not token or re.fullmatch(r'\d+', token) or re.fullmatch(r'[A-Za-z]{2,4}-\d{3,6}', token) or token.lower() in ignore:
                        continue
                    parts = re.split(r'[\s_\-]+', token)
                    city = ' '.join(p.capitalize() for p in parts if p)
                    if city:
                        return city
                return 'Ville'

            city_name = extract_city_from_filename(uploaded_file.name)
            df_feuille3['ZI'] = city_name
            df_feuille4['ZI'] = city_name

            # Helpers d'extraction complémentaires
            def extract_site_codes(messages):
                pattern = r'[A-Za-z]+-\d+'
                all_codes = []
                for message in messages:
                    codes = re.findall(pattern, message)
                    filtered = [c for c in codes if not c.upper().startswith('IMG-')]
                    all_codes.extend(filtered)
                unique_codes = []
                for c in all_codes:
                    if c not in unique_codes:
                        unique_codes.append(c)
                return unique_codes

            def extract_locations(messages):
                patterns = [
                    r'https://maps\.google\.com/q=([-\d.]+),([-\d.]+)',
                    r'localisation\s*:\s*https://maps\.google\.com/q=([-\d.]+),([-\d.]+)',
                    r'https://maps\.google\.com/\?q=([-\d.]+),([-\d.]+)',
                    r'https://maps\.google\.com/maps\?q=([-\d.]+),([-\d.]+)'
                ]
                results = []
                for msg in messages:
                    for ptn in patterns:
                        matches = re.findall(ptn, msg, re.IGNORECASE)
                        for _lat, _lon in matches:
                            results.append('OK')
                return results

            def extract_first_location_time(messages_with_time):
                patterns = [
                    r'https://maps\.google\.com/q=([-\d.]+),([-\d.]+)',
                    r'localisation\s*:\s*https://maps\.google\.com/q=([-\d.]+),([-\d.]+)',
                    r'https://maps\.google\.com/\?q=([-\d.]+),([-\d.]+)',
                    r'https://maps\.google\.com/maps\?q=([-\d.]+),([-\d.]+)'
                ]
                for md in messages_with_time:
                    for ptn in patterns:
                        if re.search(ptn, md['message'], re.IGNORECASE):
                            return md['heure']
                return None

            def extract_second_location_time(messages_with_time):
                patterns = [
                    r'https://maps\.google\.com/q=([-\d.]+),([-\d.]+)',
                    r'localisation\s*:\s*https://maps\.google\.com/q=([-\d.]+),([-\d.]+)',
                    r'https://maps\.google\.com/\?q=([-\d.]+),([-\d.]+)',
                    r'https://maps\.google\.com/maps\?q=([-\d.]+),([-\d.]+)'
                ]
                count = 0
                for md in messages_with_time:
                    for ptn in patterns:
                        if re.search(ptn, md['message'], re.IGNORECASE):
                            count += 1
                            if count == 2:
                                return md['heure']
                return None

            def extract_third_location_time(messages_with_time):
                patterns = [
                    r'https://maps\.google\.com/q=([-\d.]+),([-\d.]+)',
                    r'localisation\s*:\s*https://maps\.google\.com/q=([-\d.]+),([-\d.]+)',
                    r'https://maps\.google\.com/\?q=([-\d.]+),([-\d.]+)',
                    r'https://maps\.google\.com/maps\?q=([-\d.]+),([-\d.]+)'
                ]
                count = 0
                second = None
                third_time = None
                for md in messages_with_time:
                    for ptn in patterns:
                        if re.search(ptn, md['message'], re.IGNORECASE):
                            count += 1
                            if count == 2:
                                second = re.search(ptn, md['message'], re.IGNORECASE).group(0)
                            elif count == 3:
                                third = re.search(ptn, md['message'], re.IGNORECASE).group(0)
                                third_time = md['heure']
                                if second == third:
                                    return third_time
                                return None
                return None

            def extract_fourth_location_time(messages_with_time):
                patterns = [
                    r'https://maps\.google\.com/q=([-\d.]+),([-\d.]+)',
                    r'localisation\s*:\s*https://maps\.google\.com/q=([-\d.]+),([-\d.]+)',
                    r'https://maps\.google\.com/\?q=([-\d.]+),([-\d.]+)',
                    r'https://maps\.google\.com/maps\?q=([-\d.]+),([-\d.]+)'
                ]
                count = 0
                for md in messages_with_time:
                    for ptn in patterns:
                        if re.search(ptn, md['message'], re.IGNORECASE):
                            count += 1
                            if count == 4:
                                return md['heure']
                return None

            def extract_last_message_time(messages_with_time):
                return messages_with_time[-1]['heure'] if messages_with_time else None

            # Remplissage des feuilles 3 et 4
            for idx, r in df_unique.iterrows():
                date = r['Date']
                tech = r['Auteur']
                msgs = df[(df['Date'] == date) & (df['Auteur'] == tech)]['Message'].tolist()
                msgs_with_time = [
                    {'message': rr['Message'], 'heure': rr['Heure']}
                    for _, rr in df[(df['Date'] == date) & (df['Auteur'] == tech)].iterrows()
                ]

                site_codes = extract_site_codes(msgs)
                locations = extract_locations(msgs)
                t1 = extract_first_location_time(msgs_with_time)
                t2 = extract_second_location_time(msgs_with_time)
                t3 = extract_third_location_time(msgs_with_time)
                t4 = extract_fourth_location_time(msgs_with_time)
                t_last = extract_last_message_time(msgs_with_time)

                # Feuille 3: codes site
                idxs3 = [i for i, c in enumerate(feuille3_columns) if c == 'Code site']
                if len(site_codes) >= 2 and site_codes[0] != site_codes[1]:
                    df_feuille3.loc[idx, 'Code site'] = site_codes[0]
                    if len(idxs3) >= 2:
                        df_feuille3.iloc[idx, idxs3[1]] = site_codes[1]
                elif len(site_codes) >= 1:
                    df_feuille3.loc[idx, 'Code site'] = site_codes[0]
                    if len(idxs3) >= 2:
                        df_feuille3.iloc[idx, idxs3[1]] = 'NOK'
                else:
                    df_feuille3.loc[idx, 'Code site'] = 'NOK'
                    if len(idxs3) >= 2:
                        df_feuille3.iloc[idx, idxs3[1]] = 'NOK'

                # Feuille 3: localisations/horaires
                loc_cols3 = [i for i, c in enumerate(feuille3_columns) if c == 'Localisation']
                ok_count3 = locations.count('OK')
                for i, col_i in enumerate(loc_cols3):
                    df_feuille3.iloc[idx, col_i] = 'OK' if i < ok_count3 else 'NOK'
                dep_cols3 = [i for i, c in enumerate(feuille3_columns) if c == 'Heure de départ']
                if dep_cols3:
                    df_feuille3.iloc[idx, dep_cols3[0]] = t1 or 'NOK'
                deb_cols3 = [i for i, c in enumerate(feuille3_columns) if c == 'Heure début']
                if deb_cols3:
                    df_feuille3.iloc[idx, deb_cols3[0]] = t2 or 'NOK'
                fin_cols3 = [i for i, c in enumerate(feuille3_columns) if c == 'Heure fin']
                if fin_cols3:
                    df_feuille3.iloc[idx, fin_cols3[0]] = t3 or 'NOK'
                if len(deb_cols3) >= 2:
                    df_feuille3.iloc[idx, deb_cols3[1]] = t4 or 'NOK'
                if fin_cols3:
                    debut_s2_3 = df_feuille3.iloc[idx, deb_cols3[1]] if len(deb_cols3) >= 2 else 'NOK'
                    if debut_s2_3 == 'NOK':
                        df_feuille3.iloc[idx, fin_cols3[1]] = 'NOK'
                    elif t_last:
                        df_feuille3.iloc[idx, fin_cols3[1]] = t_last
                    else:
                        df_feuille3.iloc[idx, fin_cols3[1]] = 'NOK'

                # Feuille 3: HEURE DE TRAVAIL
                ht_cols3 = [i for i, c in enumerate(feuille3_columns) if c == 'HEURE DE TRAVAIL']
                if ht_cols3:
                    debut_s1_3 = df_feuille3.iloc[idx, deb_cols3[0]] if deb_cols3 else 'NOK'
                    fin_s2_3 = df_feuille3.iloc[idx, fin_cols3[1]] if len(fin_cols3) >= 2 else 'NOK'
                    fin_s1_3 = df_feuille3.iloc[idx, fin_cols3[0]] if fin_cols3 else 'NOK'
                    if debut_s1_3 == 'NOK':
                        df_feuille3.iloc[idx, ht_cols3[0]] = 'NOK'
                    elif fin_s2_3 != 'NOK':
                        try:
                            d = datetime.strptime(debut_s1_3, '%H:%M')
                            f = datetime.strptime(fin_s2_3, '%H:%M')
                            if f >= d:
                                diff = f - d
                                h = diff.seconds // 3600
                                m = (diff.seconds % 3600) // 60
                                df_feuille3.iloc[idx, ht_cols3[0]] = f"{h:02d}:{m:02d}"
                            else:
                                df_feuille3.iloc[idx, ht_cols3[0]] = 'NOK'
                        except Exception:
                            df_feuille3.iloc[idx, ht_cols3[0]] = 'NOK'
                    elif fin_s1_3 != 'NOK':
                        try:
                            d = datetime.strptime(debut_s1_3, '%H:%M')
                            f = datetime.strptime(fin_s1_3, '%H:%M')
                            if f >= d:
                                diff = f - d
                                h = diff.seconds // 3600
                                m = (diff.seconds % 3600) // 60
                                df_feuille3.iloc[idx, ht_cols3[0]] = f"{h:02d}:{m:02d}"
                            else:
                                df_feuille3.iloc[idx, ht_cols3[0]] = 'NOK'
                        except Exception:
                            df_feuille3.iloc[idx, ht_cols3[0]] = 'NOK'
                    else:
                        df_feuille3.iloc[idx, ht_cols3[0]] = 'NOK'

                # Feuille 4: même logique + calculs OK/NOK/SOMME/NOTE
                idxs4 = [i for i, c in enumerate(feuille4_columns) if c == 'Code site']
                if len(site_codes) >= 2 and site_codes[0] != site_codes[1]:
                    df_feuille4.loc[idx, 'Code site'] = site_codes[0]
                    if len(idxs4) >= 2:
                        df_feuille4.iloc[idx, idxs4[1]] = site_codes[1]
                elif len(site_codes) >= 1:
                    df_feuille4.loc[idx, 'Code site'] = site_codes[0]
                    if len(idxs4) >= 2:
                        df_feuille4.iloc[idx, idxs4[1]] = 'NOK'
                else:
                    df_feuille4.loc[idx, 'Code site'] = 'NOK'
                    if len(idxs4) >= 2:
                        df_feuille4.iloc[idx, idxs4[1]] = 'NOK'

                loc_cols4 = [i for i, c in enumerate(feuille4_columns) if c == 'Localisation']
                ok_count4 = locations.count('OK')
                for i, col_i in enumerate(loc_cols4):
                    df_feuille4.iloc[idx, col_i] = 'OK' if i < ok_count4 else 'NOK'
                dep_cols4 = [i for i, c in enumerate(feuille4_columns) if c == 'Heure de départ']
                if dep_cols4:
                    df_feuille4.iloc[idx, dep_cols4[0]] = t1 or 'NOK'
                deb_cols4 = [i for i, c in enumerate(feuille4_columns) if c == 'Heure début']
                if deb_cols4:
                    df_feuille4.iloc[idx, deb_cols4[0]] = t2 or 'NOK'
                fin_cols4 = [i for i, c in enumerate(feuille4_columns) if c == 'Heure fin']
                if fin_cols4:
                    df_feuille4.iloc[idx, fin_cols4[0]] = t3 or 'NOK'
                if len(deb_cols4) >= 2:
                    df_feuille4.iloc[idx, deb_cols4[1]] = t4 or 'NOK'
                if fin_cols4:
                    debut_s2_4 = df_feuille4.iloc[idx, deb_cols4[1]] if len(deb_cols4) >= 2 else 'NOK'
                    if debut_s2_4 == 'NOK':
                        df_feuille4.iloc[idx, fin_cols4[1]] = 'NOK'
                    elif t_last:
                        df_feuille4.iloc[idx, fin_cols4[1]] = t_last
                    else:
                        df_feuille4.iloc[idx, fin_cols4[1]] = 'NOK'

                ht_cols4 = [i for i, c in enumerate(feuille4_columns) if c == 'HEURE DE TRAVAIL']
                if ht_cols4:
                    debut_s1_4 = df_feuille4.iloc[idx, deb_cols4[0]] if deb_cols4 else 'NOK'
                    fin_s2_4 = df_feuille4.iloc[idx, fin_cols4[1]] if len(fin_cols4) >= 2 else 'NOK'
                    fin_s1_4 = df_feuille4.iloc[idx, fin_cols4[0]] if fin_cols4 else 'NOK'
                    if debut_s1_4 == 'NOK':
                        df_feuille4.iloc[idx, ht_cols4[0]] = 'NOK'
                    elif fin_s2_4 != 'NOK':
                        try:
                            d = datetime.strptime(debut_s1_4, '%H:%M')
                            f = datetime.strptime(fin_s2_4, '%H:%M')
                            if f >= d:
                                diff = f - d
                                h = diff.seconds // 3600
                                m = (diff.seconds % 3600) // 60
                                df_feuille4.iloc[idx, ht_cols4[0]] = f"{h:02d}:{m:02d}"
                            else:
                                df_feuille4.iloc[idx, ht_cols4[0]] = 'NOK'
                        except Exception:
                            df_feuille4.iloc[idx, ht_cols4[0]] = 'NOK'
                    elif fin_s1_4 != 'NOK':
                        try:
                            d = datetime.strptime(debut_s1_4, '%H:%M')
                            f = datetime.strptime(fin_s1_4, '%H:%M')
                            if f >= d:
                                diff = f - d
                                h = diff.seconds // 3600
                                m = (diff.seconds % 3600) // 60
                                df_feuille4.iloc[idx, ht_cols4[0]] = f"{h:02d}:{m:02d}"
                            else:
                                df_feuille4.iloc[idx, ht_cols4[0]] = 'NOK'
                        except Exception:
                            df_feuille4.iloc[idx, ht_cols4[0]] = 'NOK'
                    else:
                        df_feuille4.iloc[idx, ht_cols4[0]] = 'NOK'

                # Comptes OK/NOK/SOMME + NOTE
                loc_idx4 = [i for i, c in enumerate(feuille4_columns) if c == 'Localisation']
                code_idx4 = [i for i, c in enumerate(feuille4_columns) if c == 'Code site']
                dep_idx4 = [i for i, c in enumerate(feuille4_columns) if c == 'compteur n départ']
                indices4 = loc_idx4 + code_idx4 + dep_idx4
                ok_total = 0
                nok_total = 0
                for j in indices4:
                    v = df_feuille4.iloc[idx, j]
                    if j in dep_idx4:
                        if v == 'NOK':
                            nok_total += 1
                        else:
                            ok_total += 1
                    else:
                        if v != 'NOK' and v != '' and pd.notna(v):
                            ok_total += 1
                        else:
                            nok_total += 1
                somme = len(indices4)
                df_feuille4.loc[idx, 'NOK'] = nok_total
                df_feuille4.loc[idx, 'OK'] = ok_total
                df_feuille4.loc[idx, 'SOMME'] = somme
                df_feuille4.loc[idx, 'NOTE'] = round((ok_total / somme) * 10, 2) if somme > 0 else 'NOK'

            # Générer le fichier Excel
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_feuille1.to_excel(writer, index=False, sheet_name='Messages')
                df_feuille2.to_excel(writer, index=False, sheet_name='Messages regroupés')
                df_feuille3.to_excel(writer, index=False, sheet_name='Feuille 3')
                df_feuille4.to_excel(writer, index=False, sheet_name='Feuille 4')

                # Largeurs de colonnes de base et wrap pour la colonne Message (feuille 2)
                ws2 = writer.sheets['Messages regroupés']
                for col_idx, col_name in enumerate(df_feuille2.columns, 1):
                    if col_name == 'Message':
                        col_letter = get_column_letter(col_idx)
                        ws2.column_dimensions[col_letter].width = 80
                        for row in ws2.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                            for cell in row:
                                cell.alignment = Alignment(wrap_text=True, vertical='top')

                # Feuilles mensuelles à partir de Feuille 4
                monthly_sheet_names = []
                try:
                    df_f4_monthly = df_feuille4.copy()
                    df_f4_monthly['DATE_dt'] = pd.to_datetime(df_f4_monthly['DATE'], dayfirst=True, errors='coerce')
                    df_f4_monthly = df_f4_monthly[~df_f4_monthly['DATE_dt'].isna()].copy()
                    df_f4_monthly['year'] = df_f4_monthly['DATE_dt'].dt.year
                    df_f4_monthly['month'] = df_f4_monthly['DATE_dt'].dt.month
                    month_names_fr = {1: 'Janvier', 2: 'Février', 3: 'Mars', 4: 'Avril', 5: 'Mai', 6: 'Juin', 7: 'Juillet', 8: 'Août', 9: 'Septembre', 10: 'Octobre', 11: 'Novembre', 12: 'Décembre'}
                    for (yy, mm), grp in df_f4_monthly.groupby(['year', 'month']):
                        sheet_name = f"{month_names_fr.get(int(mm), str(int(mm)).zfill(2))} {int(yy)}"
                        grp_to_write = grp.drop(columns=['DATE_dt', 'year', 'month'], errors='ignore')
                        grp_to_write.to_excel(writer, index=False, sheet_name=sheet_name)
                        monthly_sheet_names.append(sheet_name)
                except Exception:
                    pass

                # Formatage colonnes et en-têtes colorés comme l'ancien projet
                ws3 = writer.sheets['Feuille 3']
                ws4 = writer.sheets['Feuille 4']
                for i in range(1, len(feuille3_columns) + 1):
                    ws3.column_dimensions[get_column_letter(i)].width = 20
                for i in range(1, len(feuille4_columns) + 1):
                    ws4.column_dimensions[get_column_letter(i)].width = 24

                # Couleurs
                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                light_yellow_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
                blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

                # Sections Feuille 3
                len_left = len(left_columns)
                len_driving = len(driving_headers)
                len_site1 = len(site_headers_site1)
                len_site2 = len(site_headers_site2)
                start_site1 = len_left + len_driving + 1
                end_site1 = start_site1 + len_site1 - 1
                start_site2 = end_site1 + 1
                end_site2 = start_site2 + len_site2 - 1

                # Titres fusionnés Feuille 3
                ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len_left)
                c_left = ws3.cell(row=1, column=1)
                c_left.value = "INFORMATIONS"
                c_left.fill = green_fill
                c_left.font = Font(size=16, bold=True)
                c_left.alignment = Alignment(horizontal='center', vertical='center')

                ws3.merge_cells(start_row=1, start_column=len_left + 1, end_row=1, end_column=len_left + len_driving)
                c_drv = ws3.cell(row=1, column=len_left + 1)
                c_drv.value = "DRIVING"
                c_drv.fill = yellow_fill
                c_drv.font = Font(size=16, bold=True)
                c_drv.alignment = Alignment(horizontal='center', vertical='center')

                ws3.merge_cells(start_row=1, start_column=start_site1, end_row=1, end_column=end_site1)
                c_s1 = ws3.cell(row=1, column=start_site1)
                c_s1.value = "SITE 1"
                c_s1.fill = blue_fill
                c_s1.font = Font(size=16, bold=True)
                c_s1.alignment = Alignment(horizontal='center', vertical='center')

                ws3.merge_cells(start_row=1, start_column=start_site2, end_row=1, end_column=end_site2)
                c_s2 = ws3.cell(row=1, column=start_site2)
                c_s2.value = "SITE 2"
                c_s2.fill = light_yellow_fill
                c_s2.font = Font(size=16, bold=True)
                c_s2.alignment = Alignment(horizontal='center', vertical='center')

                # Ligne d'en-têtes Feuille 3
                for idx_col, title in enumerate(feuille3_columns, start=1):
                    cell = ws3.cell(row=2, column=idx_col)
                    cell.value = title
                    cell.font = Font(size=14, bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    if 1 <= idx_col <= len_left:
                        cell.fill = green_fill
                    elif len_left < idx_col <= len_left + len_driving:
                        cell.fill = yellow_fill
                    elif start_site1 <= idx_col <= end_site1:
                        cell.fill = blue_fill
                    elif start_site2 <= idx_col <= end_site2:
                        cell.fill = light_yellow_fill

                # Lignes plus hautes
                for r in range(1, 11):
                    ws3.row_dimensions[r].height = 30

                # Feuille 4 titres fusionnés (INFORMATIONS étendu avec NOK/OK/SOMME)
                left4 = len(left_columns_feuille4)
                start_drv4 = left4 + 1
                end_drv4 = start_drv4 + len_driving - 1
                start_s1_4 = end_drv4 + 1
                end_s1_4 = start_s1_4 + len_site1 - 1
                start_s2_4 = end_s1_4 + 1
                end_s2_4 = start_s2_4 + len_site2 - 1

                ws4.merge_cells(start_row=1, start_column=1, end_row=1, end_column=left4)
                f4_left = ws4.cell(row=1, column=1)
                f4_left.value = "INFORMATIONS"
                f4_left.fill = green_fill
                f4_left.font = Font(size=16, bold=True)
                f4_left.alignment = Alignment(horizontal='center', vertical='center')

                ws4.merge_cells(start_row=1, start_column=start_drv4, end_row=1, end_column=end_drv4)
                f4_drv = ws4.cell(row=1, column=start_drv4)
                f4_drv.value = "DRIVING"
                f4_drv.fill = yellow_fill
                f4_drv.font = Font(size=16, bold=True)
                f4_drv.alignment = Alignment(horizontal='center', vertical='center')

                ws4.merge_cells(start_row=1, start_column=start_s1_4, end_row=1, end_column=end_s1_4)
                f4_s1 = ws4.cell(row=1, column=start_s1_4)
                f4_s1.value = "SITE 1"
                f4_s1.fill = blue_fill
                f4_s1.font = Font(size=16, bold=True)
                f4_s1.alignment = Alignment(horizontal='center', vertical='center')

                ws4.merge_cells(start_row=1, start_column=start_s2_4, end_row=1, end_column=end_s2_4)
                f4_s2 = ws4.cell(row=1, column=start_s2_4)
                f4_s2.value = "SITE 2"
                f4_s2.fill = light_yellow_fill
                f4_s2.font = Font(size=16, bold=True)
                f4_s2.alignment = Alignment(horizontal='center', vertical='center')

                for idx_col, title in enumerate(feuille4_columns, start=1):
                    cell = ws4.cell(row=2, column=idx_col)
                    cell.value = title
                    cell.font = Font(size=14, bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    if 1 <= idx_col <= left4:
                        cell.fill = green_fill
                    elif start_drv4 <= idx_col <= end_drv4:
                        cell.fill = yellow_fill
                    elif start_s1_4 <= idx_col <= end_s1_4:
                        cell.fill = blue_fill
                    elif start_s2_4 <= idx_col <= end_s2_4:
                        cell.fill = light_yellow_fill

                for r in range(1, 11):
                    ws4.row_dimensions[r].height = 30

                # Appliquer mêmes styles aux feuilles mensuelles
                for name in monthly_sheet_names:
                    ws_month = writer.sheets.get(name)
                    if not ws_month:
                        continue
                    # Largeurs de colonnes
                    for i in range(1, len(feuille4_columns) + 1):
                        ws_month.column_dimensions[get_column_letter(i)].width = 24
                    for r in range(1, 11):
                        ws_month.row_dimensions[r].height = 30
                    # Titres fusionnés
                    ws_month.merge_cells(start_row=1, start_column=1, end_row=1, end_column=left4)
                    m_left = ws_month.cell(row=1, column=1)
                    m_left.value = "INFORMATIONS"
                    m_left.fill = green_fill
                    m_left.font = Font(size=16, bold=True)
                    m_left.alignment = Alignment(horizontal='center', vertical='center')

                    ws_month.merge_cells(start_row=1, start_column=start_drv4, end_row=1, end_column=end_drv4)
                    m_drv = ws_month.cell(row=1, column=start_drv4)
                    m_drv.value = "DRIVING"
                    m_drv.fill = yellow_fill
                    m_drv.font = Font(size=16, bold=True)
                    m_drv.alignment = Alignment(horizontal='center', vertical='center')

                    ws_month.merge_cells(start_row=1, start_column=start_s1_4, end_row=1, end_column=end_s1_4)
                    m_s1 = ws_month.cell(row=1, column=start_s1_4)
                    m_s1.value = "SITE 1"
                    m_s1.fill = blue_fill
                    m_s1.font = Font(size=16, bold=True)
                    m_s1.alignment = Alignment(horizontal='center', vertical='center')

                    ws_month.merge_cells(start_row=1, start_column=start_s2_4, end_row=1, end_column=end_s2_4)
                    m_s2 = ws_month.cell(row=1, column=start_s2_4)
                    m_s2.value = "SITE 2"
                    m_s2.fill = light_yellow_fill
                    m_s2.font = Font(size=16, bold=True)
                    m_s2.alignment = Alignment(horizontal='center', vertical='center')

                    for idx_col, title in enumerate(feuille4_columns, start=1):
                        cell = ws_month.cell(row=2, column=idx_col)
                        cell.value = title
                        cell.font = Font(size=14, bold=True)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        if 1 <= idx_col <= left4:
                            cell.fill = green_fill
                        elif start_drv4 <= idx_col <= end_drv4:
                            cell.fill = yellow_fill
                        elif start_s1_4 <= idx_col <= end_s1_4:
                            cell.fill = blue_fill
                        elif start_s2_4 <= idx_col <= end_s2_4:
                            cell.fill = light_yellow_fill

            output.seek(0)
            filename_base = os.path.splitext(uploaded_file.name)[0]
            filename = f"perso-{filename_base}.xlsx"
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        except Exception as e:
            messages.error(request, f'Erreur lors du traitement: {str(e)}')
            return redirect('bloc3')

    return redirect('bloc3')
